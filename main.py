import re 
import time
import requests

from  colorama import Fore
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys 
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.common.exceptions import WebDriverException, NoSuchElementException


book = Workbook()
sheet = book.active
sheet['A1']="Diseño"
sheet['B1']="Marca"
sheet['C1']="Referencia"
sheet['D1']="Precio antes"
sheet['E1']="Precio despues"
count=1


def Virtual_Llantas(reference_list):
    reference_temp=""
    for reference in reference_list:
        reference_temp=reference
        reference=reference.replace('/','-')
        reference=reference.replace('lt','')
        reference=reference.replace('z','')
        reference=reference.replace('p','')
        reference=reference.replace('x','-')
        
        url=f"https://www.virtualllantas.com/llantas-{reference}"
        navegacion(url,reference_temp)
        
    #navegacion()



def list_tire_brands():
    i=2
    list_brands=[]
    file= load_workbook("tire_brands.xlsx")
    sheet_list=file.active
    while True:
        cell=sheet_list[f'A{i}'].value
        if cell!=None and not cell.isspace():
             list_brands.append(cell.lower())
             i+=1
        else:
            break
    return list_brands

def reference_list():
    i=2
    reference_list=[]
    file= load_workbook("AAAA.xlsx")
    sheet_list=file.active
    while True:
        cell=sheet_list[f'A{i}'].value
        if cell!=None and not cell.isspace():
             
             reference_list.append(cell.lower())
             i+=1
        else:
            break
    return reference_list

def Save(Detalles_llanta):
    global count
    count+=1
    for save_tire in Detalles_llanta:
        elementos=save_tire.split(',')
        count+=1
        sheet[f'A{count}']=elementos[0]
        sheet[f'B{count}']=elementos[1]
        sheet[f'C{count}']=elementos[2]
        sheet[f'D{count}']=elementos[3]
        sheet[f'E{count}']=elementos[4]
        print(count)
    book.save('prueba2_.xlsx')

def buscarLlantasValidas(llantas, lista_llantas,llantas_encontradas):
    frase_minusculas = llantas.lower()  # Convertir la frase a minúsculas para realizar una búsqueda insensible a mayúsculas y minúsculas

    for llanta in lista_llantas:
        palabra_minusculas = llanta.lower()  # Convertir la palabra a minúsculas para la comparación insensible a mayúsculas y minúsculas

        if palabra_minusculas in frase_minusculas:
            llantas_encontradas.append(llanta)
            return True
    return False

def infiniteScrollDown(browser):
    iter=1
    while True:
        time.sleep(0.35)
        scrollHeight = browser.execute_script("return document.body.scrollHeight")
        Height=250*iter
        browser.execute_script("window.scrollTo(0, " + str(Height) + ");")
        if Height > scrollHeight:
            print('End of page')
            break
        iter+=1

def virtual(browser):
    browser.find_element("xpath",'//*[@id="418535"]').click()
    browser.implicitly_wait(1)
    browser.find_element("xpath",'//*[@value="BOGOTA, D.C."]').click()
    browser.implicitly_wait(1)
    browser.find_element("xpath",'//*[@class="btn btn-primary btn-ubicacion-guardar"]').click()
    time.sleep(2)
    infiniteScrollDown(browser)
    
def navegacion(url,reference):
    browser = webdriver.Firefox()
    browser.get(url)
    browser.maximize_window()
    browser.implicitly_wait(2)

    try:
        virtual()
    except NoSuchElementException as e:
        print("Error Element :"+e)
    finally:
        virtual(browser)
    

    lista_de_llantas_validas=list_tire_brands()
    enlaces=[]
    llantas_encontradas=[]
    Detalles_llanta=[]


    html=browser.page_source

    patron = r'<a\s+href="([^"]*)"\s+title="([^"]*)"[^>]*>'
    llantas_repetida=re.findall(patron,str(html))

    browser.close()


    for llantas in llantas_repetida:
        if buscarLlantasValidas(llantas[1],lista_de_llantas_validas,llantas_encontradas):
            enlaces.append(llantas[0])
            Detalles_llanta.append(llantas[1])

    i=0
    for enlace in enlaces:
        patron_precio_antes = r'<p\s+class="antes"\s+style="[^"]+">\$<strike>([^<]+)</strike>'

        patron_precio_despues = r'<p\s+class="despues"\s+style="[^"]+">\$([^<]+)'

        resultado= requests.get(enlace)
        content= resultado.text

        precio_antes = re.search(patron_precio_antes, content)
        precio_despues = re.search(patron_precio_despues, content)

        precio_antes_valor = precio_antes.group(1) if precio_antes else None
        precio_despues_valor = precio_despues.group(1) if precio_despues else None

        
        Detalles_llanta[i]+=","+llantas_encontradas[i]
        Detalles_llanta[i]+=f",{reference}"
        Detalles_llanta[i]+=","+str(precio_antes_valor)
        Detalles_llanta[i]+=","+str(precio_despues_valor)
        
        i=i+1
    print(Detalles_llanta)
    Save(Detalles_llanta)


#(count,list)
reference_list=reference_list()

Virtual_Llantas(reference_list)














        
